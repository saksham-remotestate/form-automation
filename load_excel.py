import openpyxl
# import json


def load_evisa_excel(path: str):

    wb = openpyxl.load_workbook(path)

    ws = wb["PersonalInfo"]

    header = [cell.value for cell in ws[1]]

    values = [cell.value for cell in ws[2]]

    personal_info_row = dict(zip(header, values))

    ws2 = wb["Companions"]

    comp_header = [c.value for c in ws2[1]]
    companions = []

    for row in ws2.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        companions.append(dict(zip(comp_header, row)))

    data = {
        "step_1_citizenship": {
            "citizenship": personal_info_row["citizenship"],
            "travel_document_type": personal_info_row["travel_document_type"],
            "purpose_of_visit": personal_info_row["purpose_of_visit"],
            "insurance": bool(personal_info_row["insurance"]),
            "residence_permit": {
                "enabled": str(personal_info_row["residence_permit_enabled"]).lower(),
                "country": personal_info_row["residence_permit_country"] or ""
            }
        },
        "step_2_travel_information": {
            "arrival_date": personal_info_row["arrival_date"]
        },
        "step_4_personal_information": {
            "first_name": personal_info_row["first_name"],
            "surname": personal_info_row["surname"],
            "date_of_birth": personal_info_row["date_of_birth"],
            "sex": personal_info_row["sex"],
            "email": personal_info_row["email"],
            "travel_document_number": personal_info_row["travel_document_number"],
            "expiry_date": personal_info_row["expiry_date"],
            "representative": {
                "enabled": bool(personal_info_row["representative_enabled"]),
                "relationship": personal_info_row["representative_relationship"],
                "first_name": personal_info_row["representative_first_name"],
                "surname": personal_info_row["representative_surname"],
                "travel_document_number": personal_info_row["representative_travel_document_number"]
            },
            "companion": {
                "enabled": bool(personal_info_row["companions_enabled"]),
                "people": companions
            },
            "uploads": {
                "photo_path": personal_info_row["uploads_photo_path"],
                "passport_pdf_path": personal_info_row["uploads_passport_pdf_path"]
            }
        }
    }

    return data


# json_data = load_evisa_excel("evisa_template_v2.xlsx")
# print(json.dumps(json_data, indent=2))
